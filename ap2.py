# pip install -r requirements.txt
import requests
import sqlite3
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook

def coletar_dados_paises():
    """
    Solicita ao usuário o nome de 3 países, busca informações na API e retorna uma lista de dicionários.
    Garante que não haja países repetidos e que todos sejam válidos.
    """
    paises = []
    nomes_inseridos = set()
    i = 0
    while i < 3:
        nome = input(f"Digite o nome do país {i+1}: ").strip()
        if nome.lower() in nomes_inseridos:
            print("País já inserido, digite outro.")
            continue
        url = f"https://restcountries.com/v3.1/name/{nome}"
        try:
            resp = requests.get(url, timeout=10)
            resp.raise_for_status()
            data = resp.json()[0]
            idiomas = list(data.get('languages', {'N/A':'N/A'}).values())
            idioma_principal = idiomas[0] if idiomas else 'N/A'
            pais = {
                'nome_comum': data['name']['common'],
                'nome_oficial': data['name']['official'],
                'capital': ', '.join(data.get('capital', ['N/A'])),
                'continente': ', '.join(data.get('continents', ['N/A'])),
                'regiao': data.get('region', 'N/A'),
                'subregiao': data.get('subregion', 'N/A'),
                'populacao': data.get('population', 0),
                'area': data.get('area', 0),
                'moeda_nome': list(data.get('currencies', {'N/A':{'name':'N/A','symbol':'N/A'}}).values())[0]['name'],
                'moeda_simbolo': list(data.get('currencies', {'N/A':{'name':'N/A','symbol':'N/A'}}).values())[0]['symbol'],
                'idioma': idioma_principal,
                'fuso': ', '.join(data.get('timezones', ['N/A'])),
                'bandeira': data.get('flags', {}).get('png', 'N/A')
            }
            paises.append(pais)
            nomes_inseridos.add(nome.lower())
            i += 1
        except Exception as e:
            print(f"Erro ao obter dados para '{nome}': {e}. Tente novamente.")
    return paises

def salvar_paises_sqlite(paises):
    """
    Salva os dados dos países em um banco SQLite chamado paises.db na tabela 'paises'.
    """
    conn = sqlite3.connect('paises.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS paises (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_comum TEXT UNIQUE,
            nome_oficial TEXT,
            capital TEXT,
            continente TEXT,
            regiao TEXT,
            subregiao TEXT,
            populacao INTEGER,
            area REAL,
            moeda_nome TEXT,
            moeda_simbolo TEXT,
            idioma TEXT,
            fuso TEXT,
            bandeira TEXT
        )
    ''')
    for p in paises:
        try:
            c.execute('''
                INSERT OR IGNORE INTO paises (nome_comum, nome_oficial, capital, continente, regiao, subregiao, populacao, area, moeda_nome, moeda_simbolo, idioma, fuso, bandeira)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                p['nome_comum'], p['nome_oficial'], p['capital'], p['continente'], p['regiao'], p['subregiao'],
                p['populacao'], p['area'], p['moeda_nome'], p['moeda_simbolo'], p['idioma'], p['fuso'], p['bandeira']
            ))
        except Exception as e:
            print(f"Erro ao salvar país {p['nome_comum']}: {e}")
    conn.commit()
    conn.close()

def coletar_livros():
    """
    Realiza scraping dos 10 primeiros livros do site books.toscrape.com.
    Retorna uma lista de dicionários com título, preço, avaliação e disponibilidade.
    """
    url = "https://books.toscrape.com/"
    try:
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, 'html.parser')
        livros = []
        estrelas_map = {
            'One': 1, 'Two': 2, 'Three': 3, 'Four': 4, 'Five': 5
        }
        for artigo in soup.select('article.product_pod')[:10]:
            titulo = artigo.h3.a['title']
            preco = artigo.find('p', class_='price_color').text.strip()
            # Corrige símbolo da moeda se vier como 'Â£'
            preco = preco.replace('Â£', '£')
            estrelas = artigo.p['class'][1]  # Ex: 'Three'
            estrelas_num = estrelas_map.get(estrelas, 0)
            disponibilidade = artigo.find('p', class_='instock availability').text.strip()
            livros.append({
                'titulo': titulo,
                'preco': preco,
                'avaliacao': f"{estrelas_num} estrela(s)",
                'disponibilidade': disponibilidade
            })
        return livros
    except Exception as e:
        print(f"Erro ao coletar livros: {e}")
        return []

def salvar_livros_sqlite(livros):
    """
    Salva os dados dos livros em um banco SQLite chamado livraria.db na tabela 'livros'.
    """
    conn = sqlite3.connect('livraria.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS livros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT,
            preco TEXT,
            avaliacao TEXT,
            disponibilidade TEXT
        )
    ''')
    for l in livros:
        try:
            c.execute('''
                INSERT INTO livros (titulo, preco, avaliacao, disponibilidade)
                VALUES (?, ?, ?, ?)
            ''', (l['titulo'], l['preco'], l['avaliacao'], l['disponibilidade']))
        except Exception as e:
            print(f"Erro ao salvar livro {l['titulo']}: {e}")
    conn.commit()
    conn.close()

def ler_paises_sqlite():
    """
    Lê todos os países salvos no banco paises.db.
    """
    conn = sqlite3.connect('paises.db')
    c = conn.cursor()
    c.execute('SELECT nome_comum, nome_oficial, capital, continente, regiao, subregiao, populacao, area, moeda_nome, moeda_simbolo, idioma, fuso, bandeira FROM paises')
    paises = []
    for row in c.fetchall():
        paises.append({
            'nome_comum': row[0],
            'nome_oficial': row[1],
            'capital': row[2],
            'continente': row[3],
            'regiao': row[4],
            'subregiao': row[5],
            'populacao': row[6],
            'area': row[7],
            'moeda_nome': row[8],
            'moeda_simbolo': row[9],
            'idioma': row[10],
            'fuso': row[11],
            'bandeira': row[12]
        })
    conn.close()
    return paises

def ler_livros_sqlite():
    """
    Lê todos os livros salvos no banco livraria.db.
    """
    conn = sqlite3.connect('livraria.db')
    c = conn.cursor()
    c.execute('SELECT titulo, preco, avaliacao, disponibilidade FROM livros')
    livros = []
    for row in c.fetchall():
        livros.append({
            'titulo': row[0],
            'preco': row[1],
            'avaliacao': row[2],
            'disponibilidade': row[3]
        })
    conn.close()
    return livros

def gerar_relatorio_excel(paises, livros, nome_aluno):
    """
    Gera um relatório Excel contendo os dados dos países e livros, nome do aluno e data.
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Países"
    ws1.append(["Nome Comum", "Nome Oficial", "Capital", "Continente", "Região", "Sub-região", "População", "Área", "Moeda", "Símbolo", "Idioma Principal", "Fuso", "URL Bandeira"])
    for p in paises:
        ws1.append([
            p['nome_comum'], p['nome_oficial'], p['capital'], p['continente'], p['regiao'], p['subregiao'],
            p['populacao'], p['area'], p['moeda_nome'], p['moeda_simbolo'], p['idioma'], p['fuso'], p['bandeira']
        ])
    for col in ws1.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = max_length + 2

    ws2 = wb.create_sheet("Livros")
    ws2.append(["Título", "Preço", "Avaliação (estrelas)", "Disponibilidade"])
    for l in livros:
        ws2.append([l['titulo'], l['preco'], l['avaliacao'], l['disponibilidade']])
    for col in ws2.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = max_length + 2

    ws3 = wb.create_sheet("Relatório")
    ws3.append([f"Relatório gerado por: {nome_aluno}"])
    ws3.append([f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
    ws3.append([])
    ws3.append(["Instruções:"])
    ws3.append(["- Os dados dos países estão na aba 'Países'."])
    ws3.append(["- Os dados dos livros estão na aba 'Livros'."])
    ws3.append(["- Relatório gerado automaticamente conforme solicitado no enunciado."])
    wb.save("relatorio_final.xlsx")

# Execução do script principal
if __name__ == "__main__":
    print("=== Sistema de Extração de Dados de Países e Livros ===")
    nome_aluno = "=> JOÂO PEDRO SILVA ANTUNES - RA 2400836"
    print(f"Nome do aluno para o relatório: {nome_aluno}")
    print("\nColetando dados dos países...")
    paises = coletar_dados_paises()
    print("Salvando dados dos países no banco de dados SQLite...")
    salvar_paises_sqlite(paises)
    print("Coletando dados dos livros...")
    livros = coletar_livros()
    print("Salvando dados dos livros no banco de dados SQLite...")
    salvar_livros_sqlite(livros)
    print("Lendo dados dos bancos de dados SQLite...")
    print("Lendo dados dos países...")
    paises_db = ler_paises_sqlite()
    print("Lendo dados dos livros...")
    livros_db = ler_livros_sqlite()
    print("Gerando relatório...")
    gerar_relatorio_excel(paises_db, livros_db, nome_aluno)
    print("Processo concluído! Relatório salvo como relatorio_final.xlsx")